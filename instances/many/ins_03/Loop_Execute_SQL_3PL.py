'''
通过设置，可以周期执行SQL
'''
import logging
import os
import time
import traceback
import warnings

import openpyxl

from CLASSES.SQL_Tool import SQL_Tool
from CONFS.dev_conf import *
from instances.many.ins_03 import Sql_Control_3PL as ctr


def My_config():
    # TODO 忽略警告
    warnings.filterwarnings("ignore")

    # TODO 配置文件
    log_format = '[%(levelname)s] %(message)s [%(asctime)s]'
    date_format = '%Y-%m-%d %H:%M:%S %p'
    log_file = './LOGS/Log/file_done.log'
    logging.basicConfig(filename=log_file, filemode='a',
                        format=log_format, datefmt=date_format, level=logging.INFO)


if __name__ == '__main__':
    My_config()  # 设置基本配置

    # TODO 创建Excel
    workbook = openpyxl.Workbook()
    # TODO 保存所有sheet,k:名称 - v:对象
    sheets = {}
    # TODO 记录每个sheet已经写了多少行，便于追加
    sheet_rows = {}

    for sheetname, sql_info in ctr.execute_plain.items():
        for env, sqls in sql_info.items():
            print('正在连接%s环境……' % (env))
            times = 0
            for sql in sqls:

                try:
                    sql_tool = SQL_Tool(host=DB_ENV[env]['HOST'], port=DB_ENV[env]['PORT'], db=ctr.FINAL_DATABASE,
                                        user=DB_ENV[env]['USERNAME'], passwd=DB_ENV[env]['PASSWORD'])
                    sql_tool.connect.ping(reconnect=True)
                    times += 1  # TODO 脚本计数器
                    # TODO 执行sql
                    print('[%s] 开始执行第%i个脚本' % (env, times))
                    start = time.clock()
                    sql_tool.cursor.execute(sql)
                    endtime = time.clock()
                    print('[%s] 第%i个脚本执行完成，用时%d秒\n--------------' % (env, times, endtime - start))

                except:
                    # 将错误日志输入到目录文件中
                    f = open("./LOGS/Errors/error.log", 'a')
                    traceback.print_exc(file=f)
                    f.flush()
                    f.close()
                    print("%s环境，第%i个脚本语句查询失败，请查看日志" % (env, times))
                    exit(0)

                # TODO 判断sheet是否存在，不存在就写入Title
                if sheetname not in sheets.keys():
                    fields_info = sql_tool.cursor.description  # 获取MYSQL里面的数据字段信息
                    fields = [fields_info[i][0] for i in range(0, len(fields_info))]  # 列表生成器
                    sheet = workbook.create_sheet(title=sheetname, index=0)  # 添加工作表
                    sheets[sheetname] = sheet  # TODO sheet对象添加字典
                    sheet.cell(row=1, column=1, value='环境')  # TODO 第一行写入列名
                    for i in range(0, len(fields)):
                        sheet.cell(row=1, column=i + 2, value=fields[i])
                    sheet_rows[sheetname] = 1  # TODO sheet行数添加字典

                # TODO 写入数据
                rows = sql_tool.cursor.fetchall()
                start_row = sheet_rows[sheetname]
                sheet_rows[sheetname] = start_row + len(rows)
                for i in range(start_row, sheet_rows[sheetname]):
                    sheets[sheetname].cell(row=i + 1, column=1, value=env)
                    for j in range(0, len(fields)):
                        sheets[sheetname].cell(row=i + 1, column=j + 2, value=rows[i - start_row][j])

                del sql_tool

        # TODO 输出日志
        sheet_done = '工作表%s写入完成，共%i行数据' % (sheetname, sheet_rows[sheetname] - 1)
        print(sheet_done)
        logging.info(sheet_done)

    # TODO 保存文件
    workbook.save(os.path.join(ctr.save_path, ctr.workbook_name))
    logging.info('---------------------------')
